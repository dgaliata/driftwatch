#!/usr/bin/env python3
"""
AWS Infrastructure Scanner + Terraform Comparator
- Scans EC2, IPs, VPCs, Subnets, Security Groups, Route Tables, IGWs
- Exports to JSON, Excel, and terminal table
- Compares against terraform plan output (plan.json) for pre-deploy checks
- Compares against terraform.tfstate for drift detection
"""

import boto3
import json
import sys
import argparse
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# ── colours for terminal ────────────────────────────────────────────────────
class C:
    HEADER  = "\033[95m"
    BLUE    = "\033[94m"
    CYAN    = "\033[96m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    RED     = "\033[91m"
    BOLD    = "\033[1m"
    END     = "\033[0m"

def tag_name(tags):
    if not tags:
        return ""
    for t in tags:
        if t.get("Key") == "Name":
            return t.get("Value", "")
    return ""

def fmt_tags(tags):
    if not tags:
        return ""
    return "; ".join(f"{t['Key']}={t['Value']}" for t in tags)

def utc_now():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

# ── AWS scanning ─────────────────────────────────────────────────────────────

def scan_ec2(ec2):
    instances = []
    paginator = ec2.get_paginator("describe_instances")
    for page in paginator.paginate():
        for res in page["Reservations"]:
            for i in res["Instances"]:
                instances.append({
                    "resource_type": "ec2_instance",
                    "instance_id": i.get("InstanceId", ""),
                    "name": tag_name(i.get("Tags")),
                    "state": i.get("State", {}).get("Name", ""),
                    "instance_type": i.get("InstanceType", ""),
                    "private_ip": i.get("PrivateIpAddress", ""),
                    "public_ip": i.get("PublicIpAddress", ""),
                    "vpc_id": i.get("VpcId", ""),
                    "subnet_id": i.get("SubnetId", ""),
                    "key_name": i.get("KeyName", ""),
                    "ami_id": i.get("ImageId", ""),
                    "launch_time": str(i.get("LaunchTime", "")),
                    "security_groups": "; ".join(
                        sg.get("GroupId", "") for sg in i.get("SecurityGroups", [])
                    ),
                    "iam_profile": (i.get("IamInstanceProfile") or {}).get("Arn", ""),
                    "tags": fmt_tags(i.get("Tags")),
                })
    return instances


def scan_eips(ec2):
    addresses = []
    for addr in ec2.describe_addresses()["Addresses"]:
        addresses.append({
            "resource_type": "elastic_ip",
            "allocation_id": addr.get("AllocationId", ""),
            "public_ip": addr.get("PublicIp", ""),
            "association_id": addr.get("AssociationId", ""),
            "instance_id": addr.get("InstanceId", ""),
            "network_interface_id": addr.get("NetworkInterfaceId", ""),
            "private_ip": addr.get("PrivateIpAddress", ""),
            "domain": addr.get("Domain", ""),
            "name": tag_name(addr.get("Tags")),
            "tags": fmt_tags(addr.get("Tags")),
        })
    return addresses


def scan_vpcs(ec2):
    vpcs = []
    for v in ec2.describe_vpcs()["Vpcs"]:
        vpcs.append({
            "resource_type": "vpc",
            "vpc_id": v.get("VpcId", ""),
            "name": tag_name(v.get("Tags")),
            "cidr_block": v.get("CidrBlock", ""),
            "state": v.get("State", ""),
            "is_default": v.get("IsDefault", False),
            "dhcp_options_id": v.get("DhcpOptionsId", ""),
            "instance_tenancy": v.get("InstanceTenancy", ""),
            "tags": fmt_tags(v.get("Tags")),
        })
    return vpcs


def scan_subnets(ec2):
    subnets = []
    paginator = ec2.get_paginator("describe_subnets")
    for page in paginator.paginate():
        for s in page["Subnets"]:
            subnets.append({
                "resource_type": "subnet",
                "subnet_id": s.get("SubnetId", ""),
                "name": tag_name(s.get("Tags")),
                "vpc_id": s.get("VpcId", ""),
                "cidr_block": s.get("CidrBlock", ""),
                "availability_zone": s.get("AvailabilityZone", ""),
                "available_ips": s.get("AvailableIpAddressCount", 0),
                "map_public_ip": s.get("MapPublicIpOnLaunch", False),
                "default_for_az": s.get("DefaultForAz", False),
                "state": s.get("State", ""),
                "tags": fmt_tags(s.get("Tags")),
            })
    return subnets


def scan_security_groups(ec2):
    sgs = []
    paginator = ec2.get_paginator("describe_security_groups")
    for page in paginator.paginate():
        for sg in page["SecurityGroups"]:
            def fmt_rules(rules):
                out = []
                for r in rules:
                    proto = r.get("IpProtocol", "")
                    from_p = r.get("FromPort", "")
                    to_p = r.get("ToPort", "")
                    cidrs = [x["CidrIp"] for x in r.get("IpRanges", [])]
                    cidrs += [x["CidrIpv6"] for x in r.get("Ipv6Ranges", [])]
                    sg_src = [x["GroupId"] for x in r.get("UserIdGroupPairs", [])]
                    targets = ", ".join(cidrs + sg_src) or "all"
                    out.append(f"{proto}:{from_p}-{to_p} → {targets}")
                return "; ".join(out)

            sgs.append({
                "resource_type": "security_group",
                "group_id": sg.get("GroupId", ""),
                "group_name": sg.get("GroupName", ""),
                "description": sg.get("Description", ""),
                "vpc_id": sg.get("VpcId", ""),
                "inbound_rules": fmt_rules(sg.get("IpPermissions", [])),
                "outbound_rules": fmt_rules(sg.get("IpPermissionsEgress", [])),
                "tags": fmt_tags(sg.get("Tags")),
            })
    return sgs


def scan_route_tables(ec2):
    rts = []
    paginator = ec2.get_paginator("describe_route_tables")
    for page in paginator.paginate():
        for rt in page["RouteTables"]:
            routes = []
            for r in rt.get("Routes", []):
                dest = r.get("DestinationCidrBlock") or r.get("DestinationPrefixListId", "")
                target = (
                    r.get("GatewayId")
                    or r.get("NatGatewayId")
                    or r.get("NetworkInterfaceId")
                    or r.get("VpcPeeringConnectionId")
                    or r.get("TransitGatewayId")
                    or "local"
                )
                state = r.get("State", "")
                routes.append(f"{dest}→{target}({state})")
            assoc = "; ".join(
                a.get("SubnetId", a.get("GatewayId", "main"))
                for a in rt.get("Associations", [])
            )
            rts.append({
                "resource_type": "route_table",
                "route_table_id": rt.get("RouteTableId", ""),
                "name": tag_name(rt.get("Tags")),
                "vpc_id": rt.get("VpcId", ""),
                "routes": "; ".join(routes),
                "associations": assoc,
                "tags": fmt_tags(rt.get("Tags")),
            })
    return rts


def scan_igws(ec2):
    igws = []
    for igw in ec2.describe_internet_gateways()["InternetGateways"]:
        attached = "; ".join(
            a.get("VpcId", "") for a in igw.get("Attachments", [])
        )
        igws.append({
            "resource_type": "internet_gateway",
            "igw_id": igw.get("InternetGatewayId", ""),
            "name": tag_name(igw.get("Tags")),
            "attached_vpcs": attached,
            "state": (igw.get("Attachments") or [{}])[0].get("State", "detached"),
            "tags": fmt_tags(igw.get("Tags")),
        })
    return igws


def scan_all(region=None, profile=None):
    session_kwargs = {}
    if region:
        session_kwargs["region_name"] = region
    if profile:
        session_kwargs["profile_name"] = profile
    session = boto3.Session(**session_kwargs)
    ec2 = session.client("ec2")

    print(f"{C.CYAN}Scanning AWS infrastructure...{C.END}")
    data = {
        "scan_time": utc_now(),
        "region": ec2.meta.region_name,
        "ec2_instances": scan_ec2(ec2),
        "elastic_ips": scan_eips(ec2),
        "vpcs": scan_vpcs(ec2),
        "subnets": scan_subnets(ec2),
        "security_groups": scan_security_groups(ec2),
        "route_tables": scan_route_tables(ec2),
        "internet_gateways": scan_igws(ec2),
    }
    totals = {k: len(v) for k, v in data.items() if isinstance(v, list)}
    print(f"{C.GREEN}✓ Scan complete:{C.END} " +
          "  ".join(f"{k}={v}" for k, v in totals.items()))
    return data

# ── terminal table ────────────────────────────────────────────────────────────

def print_table(title, rows, columns):
    if not rows:
        return
    print(f"\n{C.BOLD}{C.BLUE}{'─'*6} {title} {'─'*6}{C.END}")
    widths = {c: len(c) for c in columns}
    for row in rows:
        for c in columns:
            widths[c] = max(widths[c], len(str(row.get(c, ""))))
    header = "  ".join(c.upper().ljust(widths[c]) for c in columns)
    print(f"{C.BOLD}{header}{C.END}")
    print("  ".join("─" * widths[c] for c in columns))
    for row in rows:
        print("  ".join(str(row.get(c, "")).ljust(widths[c]) for c in columns))


def print_terminal(data):
    print_table("EC2 Instances", data["ec2_instances"],
        ["instance_id","name","state","instance_type","private_ip","public_ip","vpc_id","subnet_id"])
    print_table("Elastic IPs", data["elastic_ips"],
        ["allocation_id","public_ip","instance_id","association_id","domain","name"])
    print_table("VPCs", data["vpcs"],
        ["vpc_id","name","cidr_block","state","is_default"])
    print_table("Subnets", data["subnets"],
        ["subnet_id","name","vpc_id","cidr_block","availability_zone","available_ips","map_public_ip"])
    print_table("Security Groups", data["security_groups"],
        ["group_id","group_name","vpc_id","description"])
    print_table("Route Tables", data["route_tables"],
        ["route_table_id","name","vpc_id","associations"])
    print_table("Internet Gateways", data["igws"] if "igws" in data else data.get("internet_gateways",[]),
        ["igw_id","name","attached_vpcs","state"])

# ── Excel export ──────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL      = PatternFill("solid", start_color="D6E4F0")
NORMAL_FILL   = PatternFill("solid", start_color="FFFFFF")
WARN_FILL     = PatternFill("solid", start_color="FFF2CC")
ERROR_FILL    = PatternFill("solid", start_color="FFE0E0")
GREEN_FILL    = PatternFill("solid", start_color="E2EFDA")
CELL_FONT     = Font(name="Arial", size=9)
THIN          = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row_num, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def style_data_cell(cell, row_idx, fill_override=None):
    cell.font   = CELL_FONT
    cell.border = BORDER
    cell.alignment = Alignment(wrap_text=False, vertical="center")
    if fill_override:
        cell.fill = fill_override
    else:
        cell.fill = ALT_FILL if row_idx % 2 == 0 else NORMAL_FILL

def add_sheet(wb, title, rows, columns):
    ws = wb.create_sheet(title=title)
    ws.freeze_panes = "A2"

    # header
    for ci, col in enumerate(columns, 1):
        ws.cell(row=1, column=ci, value=col.replace("_", " ").upper())
    style_header_row(ws, 1, len(columns))

    # data
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(columns, 1):
            val = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=str(val) if val != "" else "")
            style_data_cell(cell, ri)

    # auto-width (capped)
    for ci, col in enumerate(columns, 1):
        max_len = max(
            len(col),
            max((len(str(r.get(col, ""))) for r in rows), default=0)
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 50)

    ws.auto_filter.ref = ws.dimensions
    return ws


def add_summary_sheet(wb, data):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40

    # title block
    title_cell = ws["A1"]
    title_cell.value = "AWS Infrastructure Scan"
    title_cell.font  = Font(bold=True, size=16, color="1F4E79", name="Arial")
    ws.merge_cells("A1:C1")

    meta = [
        ("Scan Time", data["scan_time"]),
        ("Region",    data["region"]),
    ]
    for ri, (k, v) in enumerate(meta, 2):
        ws.cell(row=ri, column=1, value=k).font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=ri, column=2, value=v).font  = Font(name="Arial", size=10)

    # counts
    headers = ["Resource Type", "Count", "Sheet"]
    sheets_map = {
        "ec2_instances":    ("EC2 Instances",       "EC2"),
        "elastic_ips":      ("Elastic IPs",          "EIPs"),
        "vpcs":             ("VPCs",                 "VPCs"),
        "subnets":          ("Subnets",              "Subnets"),
        "security_groups":  ("Security Groups",      "SecurityGroups"),
        "route_tables":     ("Route Tables",         "RouteTables"),
        "internet_gateways":("Internet Gateways",    "IGWs"),
    }
    start_row = 5
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    for ri, (key, (label, sheet_name)) in enumerate(sheets_map.items(), start_row + 1):
        count = len(data.get(key, []))
        ws.cell(row=ri, column=1, value=label).font  = CELL_FONT
        ws.cell(row=ri, column=2, value=count).font  = CELL_FONT
        ws.cell(row=ri, column=3, value=sheet_name).font = CELL_FONT
        fill = GREEN_FILL if count > 0 else NORMAL_FILL
        for ci in range(1, 4):
            ws.cell(row=ri, column=ci).fill   = fill
            ws.cell(row=ri, column=ci).border = BORDER


def export_excel(data, path):
    if not XLSX_AVAILABLE:
        print(f"{C.YELLOW}openpyxl not installed — skipping Excel export{C.END}")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    add_summary_sheet(wb, data)
    add_sheet(wb, "EC2", data["ec2_instances"],
        ["instance_id","name","state","instance_type","private_ip","public_ip",
         "vpc_id","subnet_id","key_name","ami_id","launch_time","security_groups","iam_profile","tags"])
    add_sheet(wb, "EIPs", data["elastic_ips"],
        ["allocation_id","public_ip","association_id","instance_id",
         "network_interface_id","private_ip","domain","name","tags"])
    add_sheet(wb, "VPCs", data["vpcs"],
        ["vpc_id","name","cidr_block","state","is_default","dhcp_options_id","instance_tenancy","tags"])
    add_sheet(wb, "Subnets", data["subnets"],
        ["subnet_id","name","vpc_id","cidr_block","availability_zone",
         "available_ips","map_public_ip","default_for_az","state","tags"])
    add_sheet(wb, "SecurityGroups", data["security_groups"],
        ["group_id","group_name","vpc_id","description","inbound_rules","outbound_rules","tags"])
    add_sheet(wb, "RouteTables", data["route_tables"],
        ["route_table_id","name","vpc_id","routes","associations","tags"])
    add_sheet(wb, "IGWs", data["internet_gateways"],
        ["igw_id","name","attached_vpcs","state","tags"])

    wb.save(path)
    print(f"{C.GREEN}✓ Excel exported:{C.END} {path}")

# ── Terraform comparison ──────────────────────────────────────────────────────

def extract_tf_plan_resources(plan_path):
    """Parse `terraform show -json tfplan` output."""
    with open(plan_path) as f:
        plan = json.load(f)

    resources = []
    changes = plan.get("resource_changes", [])
    for rc in changes:
        action = rc.get("change", {}).get("actions", [])
        after  = rc.get("change", {}).get("after") or {}
        before = rc.get("change", {}).get("before") or {}
        resources.append({
            "address":       rc.get("address", ""),
            "type":          rc.get("type", ""),
            "name":          rc.get("name", ""),
            "action":        "+".join(action),
            "id_after":      after.get("id", ""),
            "id_before":     before.get("id", ""),
            "after":         after,
            "before":        before,
        })
    return resources


def extract_tf_state_resources(state_path):
    """Parse terraform.tfstate."""
    with open(state_path) as f:
        state = json.load(f)

    resources = []
    for module in state.get("resources", []):
        rtype = module.get("type", "")
        rname = module.get("name", "")
        for inst in module.get("instances", []):
            attrs = inst.get("attributes", {})
            resources.append({
                "address": f"{rtype}.{rname}",
                "type":    rtype,
                "name":    rname,
                "id":      attrs.get("id", ""),
                "attrs":   attrs,
            })
    return resources


# maps tf resource type → AWS resource type key + id field in scan data
TF_TO_SCAN_MAP = {
    "aws_instance":              ("ec2_instances",     "instance_id"),
    "aws_eip":                   ("elastic_ips",        "allocation_id"),
    "aws_vpc":                   ("vpcs",               "vpc_id"),
    "aws_subnet":                ("subnets",            "subnet_id"),
    "aws_security_group":        ("security_groups",    "group_id"),
    "aws_route_table":           ("route_tables",       "route_table_id"),
    "aws_internet_gateway":      ("internet_gateways",  "igw_id"),
}


def build_scan_index(data):
    """Build {resource_key: {id: row}} lookup from scan data."""
    idx = {}
    for key, id_field in TF_TO_SCAN_MAP.values():
        idx[key] = {r[id_field]: r for r in data.get(key, []) if r.get(id_field)}
    return idx


def compare_plan(data, plan_path):
    """Pre-deploy check: plan vs live AWS."""
    print(f"\n{C.BOLD}{C.CYAN}═══ PRE-DEPLOY CHECK (plan vs live AWS) ═══{C.END}")
    tf_resources = extract_tf_plan_resources(plan_path)
    scan_idx = build_scan_index(data)
    findings = []

    for r in tf_resources:
        rtype  = r["type"]
        action = r["action"]
        if rtype not in TF_TO_SCAN_MAP:
            continue
        scan_key, _ = TF_TO_SCAN_MAP[rtype]
        live_map = scan_idx.get(scan_key, {})
        after   = r["after"]

        finding = {
            "tf_address": r["address"],
            "action":     action,
            "status":     "ok",
            "detail":     "",
        }

        if "create" in action or "update" in action:
            # check for ID collision (update/replace might conflict)
            existing_id = after.get("id", "")
            if existing_id and existing_id in live_map:
                finding["status"] = "CONFLICT"
                finding["detail"] = f"ID {existing_id} already exists in AWS"

            # check subnet capacity for new EC2s
            if rtype == "aws_instance" and "create" in action:
                subnet_id = after.get("subnet_id", "")
                if subnet_id:
                    subnet_rows = {s["subnet_id"]: s for s in data["subnets"]}
                    if subnet_id in subnet_rows:
                        avail = subnet_rows[subnet_id].get("available_ips", 0)
                        if int(avail) < 5:
                            finding["status"] = "WARNING"
                            finding["detail"] = f"Subnet {subnet_id} only has {avail} IPs left"
                    else:
                        finding["status"] = "WARNING"
                        finding["detail"] = f"Subnet {subnet_id} not found in live scan"

            # check SG referenced by EC2 exists
            if rtype == "aws_instance":
                sg_ids = after.get("vpc_security_group_ids", []) or []
                sg_map = scan_idx.get("security_groups", {})
                missing = [s for s in sg_ids if s and s not in sg_map]
                if missing:
                    finding["status"] = "WARNING"
                    finding["detail"] = f"SGs not found in AWS: {', '.join(missing)}"

        elif "delete" in action:
            finding["status"]  = "INFO"
            finding["detail"]  = "Will be destroyed"

        findings.append(finding)

    # print results
    counts = {"ok": 0, "CONFLICT": 0, "WARNING": 0, "INFO": 0}
    for f in findings:
        s = f["status"]
        counts[s] = counts.get(s, 0) + 1
        colour = {
            "ok":       C.GREEN,
            "CONFLICT": C.RED,
            "WARNING":  C.YELLOW,
            "INFO":     C.CYAN,
        }.get(s, C.END)
        icon = {"ok": "✓", "CONFLICT": "✗", "WARNING": "⚠", "INFO": "ℹ"}.get(s, "?")
        detail = f"  {f['detail']}" if f["detail"] else ""
        print(f"  {colour}{icon} [{s}]{C.END} {f['tf_address']} ({f['action']}){detail}")

    print(f"\n  Summary — {C.GREEN}ok:{counts['ok']}{C.END}  "
          f"{C.YELLOW}warnings:{counts['WARNING']}{C.END}  "
          f"{C.RED}conflicts:{counts['CONFLICT']}{C.END}  "
          f"{C.CYAN}info:{counts['INFO']}{C.END}")
    return findings


def compare_state(data, state_path):
    """Drift detection: tfstate vs live AWS."""
    print(f"\n{C.BOLD}{C.CYAN}═══ DRIFT DETECTION (tfstate vs live AWS) ═══{C.END}")
    tf_resources = extract_tf_state_resources(state_path)
    scan_idx = build_scan_index(data)
    findings = []

    for r in tf_resources:
        rtype = r["type"]
        if rtype not in TF_TO_SCAN_MAP:
            continue
        scan_key, id_field = TF_TO_SCAN_MAP[rtype]
        live_map = scan_idx.get(scan_key, {})
        rid = r.get("id", "")

        if not rid:
            continue

        finding = {
            "tf_address": r["address"],
            "tf_id":      rid,
            "status":     "ok",
            "detail":     "",
        }

        if rid not in live_map:
            finding["status"] = "MISSING"
            finding["detail"] = f"{rtype} {rid} is in tfstate but NOT in AWS (deleted outside TF?)"
        else:
            live = live_map[rid]
            drifts = []

            # EC2 state drift
            if rtype == "aws_instance":
                tf_type  = r["attrs"].get("instance_type", "")
                live_type = live.get("instance_type", "")
                if tf_type and live_type and tf_type != live_type:
                    drifts.append(f"instance_type: TF={tf_type} AWS={live_type}")
                tf_state  = r["attrs"].get("instance_state", "")
                live_state = live.get("state", "")
                if tf_state and live_state and tf_state != live_state:
                    drifts.append(f"state: TF={tf_state} AWS={live_state}")

            # Subnet CIDR drift
            if rtype == "aws_subnet":
                tf_cidr   = r["attrs"].get("cidr_block", "")
                live_cidr = live.get("cidr_block", "")
                if tf_cidr and live_cidr and tf_cidr != live_cidr:
                    drifts.append(f"cidr_block: TF={tf_cidr} AWS={live_cidr}")

            # VPC CIDR drift
            if rtype == "aws_vpc":
                tf_cidr   = r["attrs"].get("cidr_block", "")
                live_cidr = live.get("cidr_block", "")
                if tf_cidr and live_cidr and tf_cidr != live_cidr:
                    drifts.append(f"cidr_block: TF={tf_cidr} AWS={live_cidr}")

            if drifts:
                finding["status"] = "DRIFT"
                finding["detail"] = " | ".join(drifts)

        findings.append(finding)

    # check for resources in AWS but not in state (unmanaged)
    tf_ids_by_key = {}
    for r in tf_resources:
        if r["type"] in TF_TO_SCAN_MAP:
            scan_key, _ = TF_TO_SCAN_MAP[r["type"]]
            tf_ids_by_key.setdefault(scan_key, set()).add(r.get("id", ""))

    for rtype, (scan_key, id_field) in TF_TO_SCAN_MAP.items():
        live_rows = data.get(scan_key, [])
        managed_ids = tf_ids_by_key.get(scan_key, set())
        for row in live_rows:
            rid = row.get(id_field, "")
            if rid and rid not in managed_ids:
                findings.append({
                    "tf_address": f"(unmanaged) {rtype}",
                    "tf_id":      rid,
                    "status":     "UNMANAGED",
                    "detail":     f"{rid} exists in AWS but is NOT in tfstate",
                })

    # print
    counts = {}
    for f in findings:
        s = f["status"]
        counts[s] = counts.get(s, 0) + 1
        colour = {
            "ok":        C.GREEN,
            "MISSING":   C.RED,
            "DRIFT":     C.YELLOW,
            "UNMANAGED": C.CYAN,
        }.get(s, C.END)
        icon = {"ok": "✓", "MISSING": "✗", "DRIFT": "⚠", "UNMANAGED": "ℹ"}.get(s, "?")
        detail = f"  {f['detail']}" if f["detail"] else ""
        print(f"  {colour}{icon} [{s}]{C.END} {f['tf_address']} ({f.get('tf_id','')}){detail}")

    print(f"\n  Summary — {C.GREEN}ok:{counts.get('ok',0)}{C.END}  "
          f"{C.YELLOW}drift:{counts.get('DRIFT',0)}{C.END}  "
          f"{C.RED}missing:{counts.get('MISSING',0)}{C.END}  "
          f"{C.CYAN}unmanaged:{counts.get('UNMANAGED',0)}{C.END}")
    return findings


def export_comparison_sheet(wb, findings, sheet_name, mode):
    """Add a comparison results sheet to an existing workbook."""
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"

    STATUS_FILLS = {
        "ok":        GREEN_FILL,
        "CONFLICT":  ERROR_FILL,
        "WARNING":   WARN_FILL,
        "INFO":      ALT_FILL,
        "DRIFT":     WARN_FILL,
        "MISSING":   ERROR_FILL,
        "UNMANAGED": PatternFill("solid", start_color="E8D5F5"),
    }

    if mode == "plan":
        cols = ["tf_address", "action", "status", "detail"]
    else:
        cols = ["tf_address", "tf_id", "status", "detail"]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").upper())
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    for ri, f in enumerate(findings, 2):
        fill = STATUS_FILLS.get(f.get("status", "ok"), NORMAL_FILL)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=str(f.get(col, "")))
            cell.fill   = fill
            cell.font   = CELL_FONT
            cell.border = BORDER

    for ci in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 45

# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="AWS Infrastructure Scanner + Terraform Comparator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scan only, all outputs
  python aws_infra_scan.py --output-dir ./scan_output

  # Pre-deploy check
  terraform plan -out=tfplan && terraform show -json tfplan > plan.json
  python aws_infra_scan.py --plan plan.json

  # Drift detection
  python aws_infra_scan.py --state terraform.tfstate

  # Both comparisons + specific region/profile
  python aws_infra_scan.py --plan plan.json --state terraform.tfstate \\
      --region us-west-2 --profile myprofile
        """,
    )
    parser.add_argument("--region",     help="AWS region (default: from env/config)")
    parser.add_argument("--profile",    help="AWS profile name")
    parser.add_argument("--output-dir", default=".", help="Directory for output files")
    parser.add_argument("--plan",       help="Path to terraform show -json output")
    parser.add_argument("--state",      help="Path to terraform.tfstate")
    parser.add_argument("--no-terminal",action="store_true", help="Skip terminal table output")
    parser.add_argument("--no-excel",   action="store_true", help="Skip Excel export")
    parser.add_argument("--no-json",    action="store_true", help="Skip JSON export")
    args = parser.parse_args()

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 1. scan AWS
    data = scan_all(region=args.region, profile=args.profile)

    # 2. terminal tables
    if not args.no_terminal:
        print_terminal(data)

    # 3. JSON export
    if not args.no_json:
        json_path = out_dir / f"aws_scan_{ts}.json"
        with open(json_path, "w") as f:
            json.dump(data, f, indent=2, default=str)
        print(f"{C.GREEN}✓ JSON exported:{C.END} {json_path}")

    # 4. Terraform comparisons
    plan_findings  = []
    state_findings = []

    if args.plan:
        plan_findings = compare_plan(data, args.plan)

    if args.state:
        state_findings = compare_state(data, args.state)

    # 5. Excel export (includes comparison sheets if run)
    if not args.no_excel and XLSX_AVAILABLE:
        xlsx_path = out_dir / f"aws_scan_{ts}.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        add_summary_sheet(wb, data)
        add_sheet(wb, "EC2", data["ec2_instances"],
            ["instance_id","name","state","instance_type","private_ip","public_ip",
             "vpc_id","subnet_id","key_name","ami_id","launch_time","security_groups","iam_profile","tags"])
        add_sheet(wb, "EIPs", data["elastic_ips"],
            ["allocation_id","public_ip","association_id","instance_id",
             "network_interface_id","private_ip","domain","name","tags"])
        add_sheet(wb, "VPCs", data["vpcs"],
            ["vpc_id","name","cidr_block","state","is_default","dhcp_options_id","instance_tenancy","tags"])
        add_sheet(wb, "Subnets", data["subnets"],
            ["subnet_id","name","vpc_id","cidr_block","availability_zone",
             "available_ips","map_public_ip","default_for_az","state","tags"])
        add_sheet(wb, "SecurityGroups", data["security_groups"],
            ["group_id","group_name","vpc_id","description","inbound_rules","outbound_rules","tags"])
        add_sheet(wb, "RouteTables", data["route_tables"],
            ["route_table_id","name","vpc_id","routes","associations","tags"])
        add_sheet(wb, "IGWs", data["internet_gateways"],
            ["igw_id","name","attached_vpcs","state","tags"])

        if plan_findings:
            export_comparison_sheet(wb, plan_findings, "PreDeploy_Check", "plan")
        if state_findings:
            export_comparison_sheet(wb, state_findings, "Drift_Detection", "state")

        wb.save(xlsx_path)
        print(f"{C.GREEN}✓ Excel exported:{C.END} {xlsx_path}")
    elif not args.no_excel and not XLSX_AVAILABLE:
        print(f"{C.YELLOW}⚠ openpyxl not installed. Run: pip install openpyxl{C.END}")


if __name__ == "__main__":
    main()